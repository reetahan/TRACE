from __future__ import annotations

"""
Here are the main things that can be changed manually in this code : 

- Priority calibration rates: edit DEFAULT_PRIORITY_CALIBRATION to change the share of students flagged as 
priority_student, priority_already_registered, priority_sibling, priority_parent_civil_servant, or priority_ex_student.

- Priority-school assignment weights: edit base_weights = np.array([0.55, 0.25, 0.12, 0.05, 0.02, 0.01]) 
in _choose_one_school_per_student(...) to make priority-linked schools more or less concentrated on top-ranked choices.

- Priority order / hierarchy: edit the tier assignment block if the order should differ 
from already_registered > sibling > selected priority_student > parent_civil_servant > ex_student > other.

- Tier spacing: change tier_stride=10.0 if you want larger or smaller separation between priority tiers in the score matrix.

- Lottery rule: edit the per-school lottery generation if you want a different tie-break rule than one independent random draw per school-applicant pair.

"""

import ast
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple, Union

import numpy as np
import pandas as pd
import openpyxl

DEFAULT_PRIORITY_CALIBRATION: Dict[str, float] = {
    "priority_student_student_rate": 0.6889,
    "priority_already_registered_student_rate": 0.1398,
    "priority_sibling_student_rate": 0.1539,
    "priority_parent_civil_servant_student_rate": 0.0066,
    "priority_ex_student_student_rate": 0.0521,
}


@dataclass
class PreparedChileNumbaInputs:
    student_ids: np.ndarray
    school_ids: np.ndarray
    student_rankings: List[List[int]]
    school_priority_scores: np.ndarray
    school_capacities: np.ndarray
    priority_student_seats: np.ndarray
    application_table: pd.DataFrame
    student_attributes: pd.DataFrame
    school_table: pd.DataFrame

    def to_dict(self) -> Dict[str, Any]:
        return {
            "student_ids": self.student_ids,
            "school_ids": self.school_ids,
            "student_rankings": self.student_rankings,
            "school_priority_scores": self.school_priority_scores,
            "school_capacities": self.school_capacities,
            "priority_student_seats": self.priority_student_seats,
            "application_table": self.application_table,
            "student_attributes": self.student_attributes,
            "school_table": self.school_table,
        }


# I/O helpers

def _read_excel_fast(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    data = list(rows)
    return pd.DataFrame(data, columns=header)


def _load_tabular(obj: Union[str, Path, pd.DataFrame]) -> pd.DataFrame:
    if isinstance(obj, pd.DataFrame):
        return obj.copy()
    path = Path(obj)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        return _read_excel_fast(path)
    if suffix == ".csv":
        return pd.read_csv(path)
    if suffix in {".parquet", ".pq"}:
        return pd.read_parquet(path)
    raise ValueError(f"Unsupported tabular input: {path}")


def _parse_ranking_cell(value: Any) -> List[Any]:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return []
    if isinstance(value, (list, tuple, np.ndarray, pd.Series)):
        return list(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return []
        try:
            parsed = ast.literal_eval(stripped)
            if isinstance(parsed, (list, tuple, np.ndarray)):
                return list(parsed)
        except (ValueError, SyntaxError):
            pass
        if "," in stripped:
            return [p.strip() for p in stripped.split(",") if p.strip()]
        return [stripped]
    return [value]


# Rankings conversion helper

def _rankings_to_long_df(truncated_rankings: Sequence[Sequence[str]]) -> pd.DataFrame:
    """
    Convert in-memory Mallows truncated rankings to the long application
    DataFrame format expected by the rest of the Chile pipeline.

    Each student is assigned an integer string ID ("0", "1", ...).
    Each school key is used as-is as the rbd identifier.
    """
    rows: List[Dict[str, Any]] = []
    for student_idx, ranking in enumerate(truncated_rankings):
        mrun = str(student_idx)
        for rank, prog_key in enumerate(ranking, 1):
            rows.append({
                "mrun": mrun,
                "rbd": str(prog_key),
                "preference_number": rank,
            })
    if not rows:
        return pd.DataFrame(columns=["mrun", "rbd", "preference_number"])
    return pd.DataFrame(rows)


# Application normalisation

def _normalize_applications(
    applications: Union[str, Path, pd.DataFrame],
    student_col: str = "mrun",
    school_col: str = "rbd",
    preference_col: str = "preference_number",
    ranking_col: str = "ranking",
) -> pd.DataFrame:
    df = _load_tabular(applications).copy()
    if student_col not in df.columns:
        raise ValueError(f"Missing required student column: {student_col}")

    if ranking_col in df.columns and school_col not in df.columns:
        df[ranking_col] = df[ranking_col].apply(_parse_ranking_cell)
        df = (
            df[[student_col, ranking_col]]
            .explode(ranking_col)
            .rename(columns={ranking_col: school_col})
            .dropna(subset=[school_col])
            .reset_index(drop=True)
        )
        df[preference_col] = df.groupby(student_col, sort=False).cumcount() + 1

    required = {student_col, school_col, preference_col}
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required application columns: {missing}")

    out = df[[student_col, school_col, preference_col]].copy()
    out.columns = ["mrun", "rbd", "preference_number"]
    out["mrun"] = out["mrun"].astype(str)
    out["rbd"] = out["rbd"].astype(str).str.replace(r"\.0$", "", regex=True)
    out["preference_number"] = pd.to_numeric(out["preference_number"], errors="coerce")
    out = out.dropna(subset=["preference_number"])
    out["preference_number"] = out["preference_number"].astype(int)
    out = (
        out.groupby(["mrun", "rbd"], as_index=False)["preference_number"]
        .min()
        .sort_values(["mrun", "preference_number", "rbd"], kind="stable")
        .reset_index(drop=True)
    )
    return out

# Capacity table

def _prepare_school_capacity_table(capacity_rows: Union[str, Path, pd.DataFrame]) -> pd.DataFrame:
    df = _load_tabular(capacity_rows).copy()
    required = ["rbd", "regular_seats", "priority_student_seats"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required capacity columns: {missing}")

    hs_cols = [
        c for c in [
            "high_selectivity_seats",
            "high_selectivity_seats_transitional",
            "high_selectivity_seats_ranking",
        ]
        if c in df.columns
    ]
    if "high_selectivity_seats" not in df.columns:
        df["high_selectivity_seats"] = df[hs_cols].fillna(0).sum(axis=1) if hs_cols else 0

    for col in ["regular_seats", "priority_student_seats", "high_selectivity_seats"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)


    df["rbd"] = df["rbd"].astype(str).str.replace(r"\.0$", "", regex=True)
    df["program_code"] = df["program_code"].astype(str).str.replace(r"\.0$", "", regex=True)
    df["school_id"] = df["rbd"] + "_" + df["program_code"]
    df["school_capacity"] = df["regular_seats"] + df["priority_student_seats"] + df["high_selectivity_seats"]
    return (
        df.groupby("school_id", as_index=False)[["school_capacity", "priority_student_seats"]]
        .sum()
        .sort_values("school_id", kind="stable")
        .reset_index(drop=True)
    )

# Priority attribute simulation

def _choose_one_school_per_student(
    student_school_lists: Mapping[str, Sequence[str]],
    active_students: np.ndarray,
    rng: np.random.Generator,
) -> Dict[str, str]:
    out: Dict[str, str] = {}
    base_weights = np.array([0.55, 0.25, 0.12, 0.05, 0.02, 0.01], dtype=float)

    for mrun in active_students:
        schools = list(student_school_lists.get(mrun, []))
        if schools:
            k = len(schools)

            if k <= len(base_weights):
                weights = base_weights[:k].copy()
            else:
                tail_len = k - len(base_weights)
                tail = np.full(tail_len, base_weights[-1], dtype=float)
                weights = np.concatenate([base_weights, tail])

            weights = weights / weights.sum()
            out[str(mrun)] = str(
                rng.choice(np.asarray(schools, dtype=object), p=weights)
            )
    return out


def _simulate_student_priority_attributes(
    applications_long: pd.DataFrame,
    rng: np.random.Generator,
    calibration: Optional[Mapping[str, float]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    cal = dict(DEFAULT_PRIORITY_CALIBRATION)
    if calibration is not None:
        cal.update(calibration)

    student_order = (
        applications_long[["mrun"]]
        .drop_duplicates()
        .sort_values("mrun", kind="stable")
        .reset_index(drop=True)
    )
    n_students = len(student_order)
    student_school_lists = (
        applications_long.sort_values(["mrun", "preference_number", "rbd"], kind="stable")
        .groupby("mrun")["rbd"].apply(list).to_dict()
    )

    priority_student = (rng.random(n_students) < cal["priority_student_student_rate"]).astype(np.int8)

    def active_students(rate_key: str) -> np.ndarray:
        mask = rng.random(n_students) < cal[rate_key]
        return student_order.loc[mask, "mrun"].astype(str).to_numpy()

    already_registered_school = _choose_one_school_per_student(
        student_school_lists, active_students("priority_already_registered_student_rate"), rng
    )
    sibling_school = _choose_one_school_per_student(
        student_school_lists, active_students("priority_sibling_student_rate"), rng
    )
    parent_school = _choose_one_school_per_student(
        student_school_lists, active_students("priority_parent_civil_servant_student_rate"), rng
    )
    ex_school = _choose_one_school_per_student(
        student_school_lists, active_students("priority_ex_student_student_rate"), rng
    )

    student_attributes = student_order.copy()
    student_attributes["priority_student"] = priority_student
    student_attributes["priority_already_registered_school"] = student_attributes["mrun"].map(already_registered_school)
    student_attributes["priority_sibling_school"] = student_attributes["mrun"].map(sibling_school)
    student_attributes["priority_parent_civil_servant_school"] = student_attributes["mrun"].map(parent_school)
    student_attributes["priority_ex_student_school"] = student_attributes["mrun"].map(ex_school)

    app = applications_long.copy()
    student_priority_map = dict(zip(student_attributes["mrun"], student_attributes["priority_student"]))
    app["priority_student"] = app["mrun"].map(student_priority_map).fillna(0).astype(np.int8)
    app["priority_already_registered"] = (
        app["mrun"].map(already_registered_school).fillna("__none__") == app["rbd"]
    ).astype(np.int8)
    app["priority_sibling"] = (
        app["mrun"].map(sibling_school).fillna("__none__") == app["rbd"]
    ).astype(np.int8)
    app["priority_parent_civil_servant"] = (
        app["mrun"].map(parent_school).fillna("__none__") == app["rbd"]
    ).astype(np.int8)
    app["priority_ex_student"] = (
        app["mrun"].map(ex_school).fillna("__none__") == app["rbd"]
    ).astype(np.int8)
    return app, student_attributes

# Priority tiers + dense score matrix

def _assign_school_level_priority_tiers_and_dense_scores(
    applications_long: pd.DataFrame,
    school_table: pd.DataFrame,
    rng: np.random.Generator,
    tier_stride: float = 10.0,
    student_lottery=None
):
    school_ids = school_table["school_id"].astype(str).tolist()
    student_ids = sorted(applications_long["mrun"].astype(str).unique().tolist())
    school_to_idx = {s: i for i, s in enumerate(school_ids)}
    student_to_idx = {s: i for i, s in enumerate(student_ids)}

    filtered = applications_long[applications_long["rbd"].isin(school_to_idx)].copy()
    if filtered.empty:
        raise ValueError("No application rows remain after intersecting with capacity schools.")

    filtered["student_idx"] = filtered["mrun"].map(student_to_idx).astype(np.int32)
    filtered["school_idx"] = filtered["rbd"].map(school_to_idx).astype(np.int32)
    filtered = (
        filtered
        .sort_values(["student_idx", "preference_number", "school_idx"], kind="stable")
        .reset_index(drop=True)
    )

    n_students = len(student_ids)
    n_schools = len(school_ids)

    if student_lottery is not None:
        filtered["lottery"] = filtered["mrun"].astype(str).map(student_lottery)
    else:
        lottery_col = np.empty(len(filtered), dtype=np.float64)
        for school_idx_val, grp in filtered.groupby("school_idx", sort=True):
            lottery_col[grp.index] = rng.random(len(grp))
        filtered["lottery"] = lottery_col


    school_table_idx = school_table.set_index("school_id")
    capacities = school_table_idx.loc[school_ids, "school_capacity"].to_numpy(dtype=np.int32)
    priority_seats = school_table_idx.loc[school_ids, "priority_student_seats"].to_numpy(dtype=np.int32)

    quota_caps = np.minimum(priority_seats, capacities)
    eligible = filtered[
        (filtered["priority_student"] == 1)
        & (filtered["priority_already_registered"] == 0)
        & (filtered["priority_sibling"] == 0)
    ][["student_idx", "school_idx", "lottery"]].copy()

    if eligible.empty:
        selected_pairs: set[tuple[int, int]] = set()
    else:
        eligible = eligible.sort_values(["school_idx", "lottery", "student_idx"], kind="stable")
        eligible["quota_rank"] = eligible.groupby("school_idx").cumcount()
        school_quota_map = pd.Series(quota_caps, index=np.arange(n_schools))
        eligible["quota_cap"] = eligible["school_idx"].map(school_quota_map).astype(np.int32)
        selected_rows = eligible[eligible["quota_rank"] < eligible["quota_cap"]]
        selected_pairs = set(
            zip(selected_rows["student_idx"].astype(int), selected_rows["school_idx"].astype(int))
        )

    student_idx_arr = filtered["student_idx"].to_numpy(dtype=np.int32)
    school_idx_arr = filtered["school_idx"].to_numpy(dtype=np.int32)
    already_arr = filtered["priority_already_registered"].to_numpy(dtype=np.int8)
    sibling_arr = filtered["priority_sibling"].to_numpy(dtype=np.int8)
    parent_arr = filtered["priority_parent_civil_servant"].to_numpy(dtype=np.int8)
    ex_arr = filtered["priority_ex_student"].to_numpy(dtype=np.int8)

    selected_mask = np.fromiter(
        ((int(st), int(sc)) in selected_pairs for st, sc in zip(student_idx_arr, school_idx_arr)),
        dtype=np.bool_,
        count=len(filtered),
    )

    tiers = np.full(len(filtered), 6, dtype=np.int8)
    reasons = np.full(len(filtered), "other", dtype=object)

    mask = already_arr == 1
    tiers[mask] = 1
    reasons[mask] = "priority_already_registered"

    mask = (tiers == 6) & (sibling_arr == 1)
    tiers[mask] = 2
    reasons[mask] = "priority_sibling"

    mask = (tiers == 6) & selected_mask
    tiers[mask] = 3
    reasons[mask] = "priority_student_selected_by_quota"

    mask = (tiers == 6) & (parent_arr == 1)
    tiers[mask] = 4
    reasons[mask] = "priority_parent_civil_servant"

    mask = (tiers == 6) & (ex_arr == 1)
    tiers[mask] = 5
    reasons[mask] = "priority_ex_student"

    filtered["priority_tier"] = tiers
    filtered["priority_reason"] = reasons
    filtered["selected_priority_student"] = selected_mask.astype(np.int8)

    tie_eps = (student_idx_arr.astype(np.float64) + 1.0) * 1e-12
    scores = (
        tiers.astype(np.float64) * tier_stride
        + filtered["lottery"].to_numpy(dtype=np.float64)
        + tie_eps
    )
    filtered["numba_score"] = scores

    sentinel = 9999.0 + tier_stride * 7
    dense_scores = np.full((n_schools, n_students), sentinel, dtype=np.float64)
    dense_scores[school_idx_arr, student_idx_arr] = scores

    rank_df = (
        filtered[["student_idx", "school_idx", "preference_number"]]
        .sort_values(["student_idx", "preference_number", "school_idx"], kind="stable")
        .drop_duplicates(subset=["student_idx", "school_idx"], keep="first")
    )
    student_rankings: List[List[int]] = [[] for _ in range(n_students)]
    for row in rank_df.itertuples(index=False):
        student_rankings[int(row.student_idx)].append(int(row.school_idx))

    return (
        np.asarray(student_ids, dtype=object),
        student_rankings,
        dense_scores,
        filtered.reset_index(drop=True),
        capacities,
        priority_seats,
        np.asarray(school_ids, dtype=object),
    )

# API

def prepare_chile_numba_inputs(
    applications: Union[str, Path, pd.DataFrame],
    capacity_rows: Union[str, Path, pd.DataFrame],
    *,
    seed: int = 42,
    calibration: Optional[Mapping[str, float]] = None,
    student_col: str = "mrun",
    school_col: str = "rbd",
    preference_col: str = "preference_number",
    ranking_col: str = "ranking",
    student_lottery=None
) -> Dict[str, Any]:
    rng = np.random.default_rng(seed)
    applications_long = _normalize_applications(
        applications,
        student_col=student_col,
        school_col=school_col,
        preference_col=preference_col,
        ranking_col=ranking_col,
    )
    school_table = _prepare_school_capacity_table(capacity_rows)
    app_with_flags, student_attributes = _simulate_student_priority_attributes(
        applications_long, rng, calibration
    )
    (
        student_ids,
        student_rankings,
        school_priority_scores,
        application_table,
        school_capacities,
        priority_student_seats,
        school_ids,
    ) = _assign_school_level_priority_tiers_and_dense_scores(app_with_flags, school_table, rng, student_lottery=student_lottery)
    return PreparedChileNumbaInputs(
        student_ids=student_ids,
        school_ids=school_ids,
        student_rankings=student_rankings,
        school_priority_scores=school_priority_scores,
        school_capacities=school_capacities,
        priority_student_seats=priority_student_seats,
        application_table=application_table,
        student_attributes=student_attributes,
        school_table=school_table,
    ).to_dict()


def prepare_chile_numba_inputs_from_rankings(
    truncated_rankings: Sequence[Sequence[str]],
    capacity_rows: Union[str, Path, pd.DataFrame],
    *,
    seed: int = 42,
    calibration: Optional[Mapping[str, float]] = None,
) -> Dict[str, Any]:
    """
    Alternative entry point that accepts in-memory Mallows truncated rankings
    instead of a CSV/Excel application file.

    Expected input:
    - truncated_rankings: list of lists of school/program keys, one per student,
      already truncated (e.g. output of list_length.py in the NYC pipeline).
    - capacity_rows: same capacity table as prepare_chile_numba_inputs.
    - seed, calibration: same as prepare_chile_numba_inputs.

    Student IDs are assigned as "0", "1", ... in the order of truncated_rankings.
    The rest of the pipeline (priority simulation, score matrix, GS compatibility)
    is identical to prepare_chile_numba_inputs.
    """
    rng = np.random.default_rng(seed)
    applications_long = _rankings_to_long_df(truncated_rankings)
    if applications_long.empty:
        raise ValueError("truncated_rankings is empty or contains no school keys.")
    school_table = _prepare_school_capacity_table(capacity_rows)
    app_with_flags, student_attributes = _simulate_student_priority_attributes(
        applications_long, rng, calibration
    )
    (
        student_ids,
        student_rankings,
        school_priority_scores,
        application_table,
        school_capacities,
        priority_student_seats,
        school_ids,
    ) = _assign_school_level_priority_tiers_and_dense_scores(app_with_flags, school_table, rng)
    return PreparedChileNumbaInputs(
        student_ids=student_ids,
        school_ids=school_ids,
        student_rankings=student_rankings,
        school_priority_scores=school_priority_scores,
        school_capacities=school_capacities,
        priority_student_seats=priority_student_seats,
        application_table=application_table,
        student_attributes=student_attributes,
        school_table=school_table,
    ).to_dict()


def save_prepared_inputs(prepared: Mapping[str, Any], output_dir: Union[str, Path]) -> None:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(
        output_dir / "numba_inputs.npz",
        school_priority_scores=prepared["school_priority_scores"],
        school_capacities=prepared["school_capacities"],
        priority_student_seats=prepared["priority_student_seats"],
        student_ids=prepared["student_ids"],
        school_ids=prepared["school_ids"],
    )
    prepared["application_table"].to_csv(output_dir / "application_table.csv", index=False)
    prepared["student_attributes"].to_csv(output_dir / "student_attributes.csv", index=False)
    prepared["school_table"].to_csv(output_dir / "school_table.csv", index=False)
    ranking_df = pd.DataFrame(
        {
            "student_id": prepared["student_ids"],
            "student_rankings": [json.dumps(r) for r in prepared["student_rankings"]],
        }
    )
    ranking_df.to_csv(output_dir / "student_rankings.csv", index=False)
    meta = {
        "n_students": int(len(prepared["student_ids"])),
        "n_schools": int(len(prepared["school_ids"])),
        "n_application_rows": int(len(prepared["application_table"])),
        "notes": (
            "Use student_rankings + school_priority_scores + school_capacities "
            "with EduRanker gale_shapley_per_school_numba_wrapper. "
            "Lottery is per-school MTB: each school draws independently for its own applicants. "
            "padded_rankings and ranking_lengths are intentionally omitted because the wrapper rebuilds them."
        ),
    }
    with open(output_dir / "metadata.json", "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2)

# Main

def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(
        description="Build Chile-style Numba inputs for EduRanker Gale-Shapley."
    )
    parser.add_argument("--applications", required=True)
    parser.add_argument("--capacity", required=True)
    parser.add_argument("--output_dir", default=None)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()
    prepared = prepare_chile_numba_inputs(args.applications, args.capacity, seed=args.seed)
    if args.output_dir:
        save_prepared_inputs(prepared, args.output_dir)
        print(f"Saved outputs to: {args.output_dir}")
    print(f"Students: {len(prepared['student_ids']):,}")
    print(f"Schools:  {len(prepared['school_ids']):,}")
    print(f"Application rows: {len(prepared['application_table']):,}")
    print(f"Priority score matrix shape: {prepared['school_priority_scores'].shape}")
    example_lengths = [len(r) for r in prepared["student_rankings"][:5]]
    print(f"Example ranking lengths (first 5 students): {example_lengths}")


if __name__ == "__main__":
    main()
